"""Exporta o template .xlsx preenchido a partir de um WorkbookData (loader)."""

from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

import config
import insights
import loader

SHEET_RELATORIO = config.SHEET_RELATORIO
SHEET_GRAFICOS = config.SHEET_GRAFICOS
SHEET_INSIGHTS = "INSIGHTS"


def _data_br(valor: datetime | None) -> str | None:
    if valor is None:
        return None
    return valor.strftime("%d/%m/%Y")


def _agrupar_categorias(local: loader.Local) -> list[tuple[str, list[loader.Item]]]:
    ordem: list[str] = []
    grupos: dict[str, list[loader.Item]] = {}
    for item in local.itens:
        categoria = item.categoria or "GERAL"
        if categoria not in grupos:
            grupos[categoria] = []
            ordem.append(categoria)
        grupos[categoria].append(item)
    return [(c, grupos[c]) for c in ordem]


def _montar_aba_local(ws, local: loader.Local) -> None:
    linha = 1
    for categoria, itens in _agrupar_categorias(local):
        ws.cell(row=linha, column=2, value=f"{config.PREFIXO_MATERIAL}{categoria}")
        linha += 1
        for item in itens:
            ws.cell(row=linha, column=1, value=item.cod)
            ws.cell(row=linha, column=2, value=item.material)
            ws.cell(row=linha, column=3, value=item.qtd)
            ws.cell(row=linha, column=4, value=item.valor_unit)
            ws.cell(row=linha, column=5, value=item.valor_total)
            linha += 1
        ws.cell(row=linha, column=1, value=config.HEADER_TOTAL)
        linha += 1


def _montar_aba_insights(wb: openpyxl.Workbook, workbook: loader.WorkbookData) -> None:
    ws = wb.create_sheet(SHEET_INSIGHTS)
    ws.append(["LOCAL", "SEVERIDADE", "INSIGHT"])
    for local in workbook.locais:
        for insight in insights.gerar_insights(local):
            ws.append([local.nome, insight["severidade"], insight["texto"]])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="172033")
        cell.alignment = Alignment(horizontal="center")
    for linha in ws.iter_rows(min_row=2):
        linha[2].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 110
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def montar_planilha(workbook: loader.WorkbookData) -> BytesIO:
    wb = openpyxl.Workbook()
    ws_rel = wb.active
    ws_rel.title = SHEET_RELATORIO

    for local in workbook.locais:
        aba = local.nome or "Local"
        wb.create_sheet(aba[:31])
        _montar_aba_local(wb[aba[:31]], local)

    cabecalho = [
        "LOCAL", "VALOR MENSAL", "TAXA INSTALACAO", "IMPOSTOS", "SALDO APOS",
        "CUSTO MANUT", "TERCEIRIZADA", "CHIP", "SOFTWARES", "SALDO MENSAL",
        "MAO DE OBRA", "EQUIPAMENTO", "INVESTIMENTO", "TEMPO RETORNO", "DATA INST",
    ]
    ws_rel.append(cabecalho)
    for local in workbook.locais:
        ws_rel.append(
            [
                local.nome,
                local.valor_mensal,
                local.taxa_instalacao,
                round(local.impostos, 2),
                round(local.saldo_apos_impostos, 2),
                local.custo_manutencao,
                local.mensal_terceirizada,
                local.chip_mensal,
                local.custos_softwares,
                round(local.saldo_mensal, 2),
                local.mao_de_obra,
                round(local.equipamento, 2),
                round(local.investimento, 2),
                round(local.tempo_retorno, 2) if local.tempo_retorno is not None else None,
                _data_br(local.data_inst),
            ]
        )

    _montar_aba_insights(wb, workbook)
    wb.create_sheet(SHEET_GRAFICOS)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
