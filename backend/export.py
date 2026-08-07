from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import analysis


def _estilo_cabecalho(ws, linha: int, n_colunas: int) -> None:
    preenchimento = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fonte = Font(color="FFFFFF", bold=True, size=11)
    for coluna in range(1, n_colunas + 1):
        celula = ws.cell(row=linha, column=coluna)
        celula.fill = preenchimento
        celula.font = fonte
    ws.freeze_panes = f"A{linha + 1}"


def exportar_excel(locais) -> BytesIO:
    wb = openpyxl.Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Resumo por local"
    colunas_resumo = [
        "Local",
        "Receita mensal",
        "Receita anual",
        "Impostos (15%)",
        "Saldo mensal",
        "Margem",
        "Mão de obra",
        "Equipamento",
        "Investimento",
        "Tempo de retorno (meses)",
        "Itens",
        "Instalação",
    ]
    ws_resumo.append(colunas_resumo)
    _estilo_cabecalho(ws_resumo, 1, len(colunas_resumo))
    for local in locais:
        res = analysis.resumo(local)
        ws_resumo.append(
            [
                res["local"],
                res["valor_mensal"],
                res["receita_anual"],
                res["impostos"],
                res["saldo_mensal"],
                res["margem"] if res["margem"] is not None else "",
                res["mao_de_obra"],
                res["equipamento"],
                res["investimento"],
                res["tempo_retorno"] if res["tempo_retorno"] is not None else "",
                res["num_itens"],
                res["data_inst"].strftime("%d/%m/%Y") if res["data_inst"] else "",
            ]
        )
    for coluna in range(2, len(colunas_resumo) + 1):
        ws_resumo.cell(row=1, column=coluna).number_format = "#,##0.00"
    ws_resumo.cell(row=1, column=6).number_format = "0.00%"
    ws_resumo.column_dimensions["A"].width = 34
    for coluna in range(2, len(colunas_resumo) + 1):
        ws_resumo.column_dimensions[get_column_letter(coluna)].width = 18

    ws_itens = wb.create_sheet("Itens por local")
    colunas_itens = ["Local", "Categoria", "Código", "Material", "Qtd", "Valor unit.", "Valor total"]
    ws_itens.append(colunas_itens)
    _estilo_cabecalho(ws_itens, 1, len(colunas_itens))
    for local in locais:
        for item in sorted(local.itens, key=lambda i: i.valor_total, reverse=True):
            ws_itens.append(
                [
                    local.nome,
                    item.categoria,
                    item.cod,
                    item.material,
                    item.qtd,
                    item.valor_unit,
                    item.valor_total,
                ]
            )
    ws_itens.column_dimensions["A"].width = 34
    ws_itens.column_dimensions["B"].width = 20
    ws_itens.column_dimensions["C"].width = 12
    ws_itens.column_dimensions["D"].width = 52
    ws_itens.column_dimensions["E"].width = 10
    ws_itens.column_dimensions["F"].width = 14
    ws_itens.column_dimensions["G"].width = 14

    ws_comp = wb.create_sheet("Comparativo do projeto")
    resumos = [analysis.resumo(local) for local in locais]
    totais = {
        "receita_mensal": sum(r["valor_mensal"] for r in resumos),
        "receita_anual": sum(r["receita_anual"] for r in resumos),
        "saldo_mensal": sum(r["saldo_mensal"] for r in resumos),
        "investimento": sum(r["investimento"] for r in resumos),
        "equipamento": sum(r["equipamento"] for r in resumos),
        "mao_de_obra": sum(r["mao_de_obra"] for r in resumos),
        "num_locais": len(resumos),
        "num_itens": sum(r["num_itens"] for r in resumos),
    }
    ws_comp.append(["Métrica", "Total do projeto"])
    _estilo_cabecalho(ws_comp, 1, 2)
    for rotulo, campo in [
        ("Locais", "num_locais"),
        ("Itens", "num_itens"),
        ("Receita mensal", "receita_mensal"),
        ("Receita anual", "receita_anual"),
        ("Saldo mensal", "saldo_mensal"),
        ("Equipamento", "equipamento"),
        ("Mão de obra", "mao_de_obra"),
        ("Investimento", "investimento"),
    ]:
        ws_comp.append([rotulo, totais[campo]])
    ws_comp.column_dimensions["A"].width = 24
    ws_comp.column_dimensions["B"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
