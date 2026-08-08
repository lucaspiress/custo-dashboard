from io import BytesIO
from pathlib import Path

import openpyxl

SHEET_RELATORIO = "RELATORIO"
SHEET_GRAFICOS = "GRÁFICOS"


def _montar_aba_local(ws, categoria: str, itens: list[dict]) -> None:
    ws.append([None, f"MATERIAL {categoria}"])
    for item in itens:
        ws.append(
            [
                item["cod"],
                item["material"],
                item["qtd"],
                item["valor_unit"],
                item["qtd"] * item["valor_unit"],
            ]
        )
    ws.append(["TOTAL"])


def montar_planilha_teste() -> BytesIO:
    wb = openpyxl.Workbook()
    ws_rel = wb.active
    ws_rel.title = SHEET_RELATORIO

    locais = [
        {
            "nome": "SESC TESTE",
            "valor_mensal": 10000,
            "mao_de_obra": 2000,
            "itens": [
                {"categoria": "ALARME", "itens": [
                    {"cod": "A1", "material": "Central de alarme", "qtd": 1, "valor_unit": 1500},
                    {"cod": "A2", "material": "Sensor de presença", "qtd": 8, "valor_unit": 120},
                ]},
                {"categoria": "CFTV", "itens": [
                    {"cod": "C1", "material": "Câmera IP 2MP", "qtd": 4, "valor_unit": 850},
                    {"cod": "C2", "material": "DVR 8 canais", "qtd": 1, "valor_unit": 1400},
                ]},
            ],
        },
        {
            "nome": "UNIDADE B",
            "valor_mensal": 6000,
            "mao_de_obra": 1200,
            "itens": [
                {"categoria": "CFTV", "itens": [
                    {"cod": "C1", "material": "Câmera IP 2MP", "qtd": 2, "valor_unit": 850},
                ]},
            ],
        },
    ]

    for local in locais:
        nome = local["nome"]
        wb.create_sheet(nome)
        ws_local = wb[nome]
        for bloco in local["itens"]:
            _montar_aba_local(ws_local, bloco["categoria"], bloco["itens"])

    ws_rel.append(["LOCAL", "VALOR MENSAL", "TAXA INSTALACAO", "IMPOSTOS", "SALDO APOS",
                   "CUSTO MANUT", "TERCEIRIZADA", "CHIP", "SOFTWARES", "SALDO MENSAL",
                   "MAO DE OBRA", "EQUIPAMENTO", "INVESTIMENTO", "TEMPO RETORNO", "DATA INST"])
    for linha, local in enumerate(locais, start=2):
        nome = local["nome"]
        equipamento = sum(
            item["qtd"] * item["valor_unit"]
            for bloco in local.get("itens", [])
            for item in bloco["itens"]
        )
        ws_rel.cell(row=linha, column=1, value=nome)
        ws_rel.cell(row=linha, column=2, value=local.get("valor_mensal", 5000))
        ws_rel.cell(row=linha, column=3, value=local.get("taxa_instalacao", 0))
        ws_rel.cell(row=linha, column=6, value=local.get("custo_manutencao", 0))
        ws_rel.cell(row=linha, column=7, value=local.get("mensal_terceirizada", 0))
        ws_rel.cell(row=linha, column=8, value=local.get("chip_mensal", 0))
        ws_rel.cell(row=linha, column=9, value=local.get("custos_softwares", 0))
        ws_rel.cell(row=linha, column=11, value=local.get("mao_de_obra", 1000))
        ws_rel.cell(row=linha, column=12, value=f"='{nome}'!E99")
        ws_rel.cell(row=linha, column=13, value=local.get("mao_de_obra", 1000) + equipamento)
        ws_rel.cell(row=linha, column=15, value=local.get("data_inst", "01/08/2026"))

    wb.create_sheet(SHEET_GRAFICOS)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def salvar_planilha_teste(destino: Path) -> Path:
    destino.write_bytes(montar_planilha_teste().getvalue())
    return destino
