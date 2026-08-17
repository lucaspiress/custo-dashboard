from io import BytesIO

import openpyxl

from fixtures import planilha_base


def _login(cliente, username="admin", senha="admin123456"):
    resposta = cliente.post("/api/auth/login", json={"username": username, "senha": senha})
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def _criar_projeto(cliente, nome="Projeto Teste", cliente_nome=None):
    corpo = {"nome": nome}
    if cliente_nome:
        corpo["cliente"] = cliente_nome
    resposta = cliente.post("/api/projetos", json=corpo)
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def _criar_local(cliente, projeto_id, **campos):
    dados = {"nome": "SESC TESTE", "valor_mensal": 10000, "mao_de_obra": 2000}
    dados.update(campos)
    resposta = cliente.post(f"/api/projetos/{projeto_id}/locais", json=dados)
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def _criar_item(cliente, local_id, **campos):
    dados = {"categoria": "ALARME", "material": "Central de alarme", "qtd": 1, "valor_unit": 1500}
    dados.update(campos)
    resposta = cliente.post(f"/api/projetos/locais/{local_id}/itens", json=dados)
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def test_health(cliente):
    resposta = cliente.get("/api/health")
    assert resposta.status_code == 200
    assert resposta.json()["ok"] is True


def test_login_errado(cliente):
    resposta = cliente.post("/api/auth/login", json={"username": "admin", "senha": "senha-errada"})
    assert resposta.status_code == 401


def test_me_sem_sessao(cliente):
    assert cliente.get("/api/auth/me").status_code == 401


def test_projetos_exige_login(cliente):
    assert cliente.get("/api/projetos").status_code == 401
    assert cliente.post("/api/projetos", json={"nome": "X"}).status_code == 401
    assert cliente.post("/api/projetos/importar").status_code == 401


def test_criar_e_listar_projeto(cliente, admin):
    projeto = _criar_projeto(cliente, nome="Cliente A", cliente_nome="ACME")
    assert projeto["nome"] == "Cliente A"
    assert projeto["cliente"] == "ACME"
    lista = cliente.get("/api/projetos").json()
    assert len(lista) == 1
    assert lista[0]["id"] == projeto["id"]
    assert lista[0]["num_locais"] == 0
    assert lista[0]["totais"]["investimento"] == 0


def test_criar_projeto_sem_nome(cliente, admin):
    resposta = cliente.post("/api/projetos", json={"nome": "  "})
    assert resposta.status_code == 400


def test_renomear_projeto(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.patch(f"/api/projetos/{projeto['id']}", json={"nome": "Novo Nome", "cliente": "Rota"})
    assert resposta.status_code == 200
    assert resposta.json()["nome"] == "Novo Nome"
    assert resposta.json()["cliente"] == "Rota"


def test_excluir_projeto_cascata(cliente, admin):
    projeto = _criar_projeto(cliente)
    local = _criar_local(cliente, projeto["id"])
    _criar_item(cliente, local["id"])
    assert cliente.delete(f"/api/projetos/{projeto['id']}").status_code == 204
    assert cliente.get("/api/projetos").json() == []
    assert cliente.get(f"/api/projetos/{projeto['id']}").status_code == 404


def test_dashboard_do_projeto_calcula(cliente, admin):
    projeto = _criar_projeto(cliente)
    local = _criar_local(cliente, projeto["id"])
    _criar_item(cliente, local["id"], cod="A1")
    _criar_item(cliente, local["id"], material="Sensor de presença", qtd=8, valor_unit=120, cod="A2")
    _criar_item(cliente, local["id"], categoria="CFTV", material="Câmera IP 2MP", qtd=4, valor_unit=850, cod="C1")
    _criar_item(cliente, local["id"], categoria="CFTV", material="DVR 8 canais", qtd=1, valor_unit=1400, cod="C2")

    dados = cliente.get(f"/api/projetos/{projeto['id']}").json()
    assert dados["filename"] == "Projeto Teste"
    assert len(dados["locais"]) == 1
    sesc = dados["locais"][0]
    assert sesc["resumo"]["valor_mensal"] == 10000
    assert sesc["resumo"]["impostos"] == 1500
    assert sesc["resumo"]["saldo_mensal"] == 8500
    equip_esperado = 1500 + 8 * 120 + 4 * 850 + 1400
    assert sesc["resumo"]["equipamento"] == equip_esperado
    assert sesc["resumo"]["investimento"] == 2000 + equip_esperado
    assert len(sesc["itens"]) == 4
    assert len(sesc["insights"]) > 0
    assert sesc["graficos"]["pareto"].startswith("{")
    assert sesc["graficos"]["payback"].startswith("{")

    fluxo = sesc["fluxo"]["24"]
    assert fluxo["meses"] == 24
    assert len(fluxo["pontos"]) == 24
    assert fluxo["grafico"].startswith("{")
    assert len({l["fluxo"][h]["meses"] for l in dados["locais"] for h in ("6", "12", "24", "36")}) == 4

    projeto_resumo = dados["projeto"]
    assert projeto_resumo["totais"]["num_locais"] == 1
    assert projeto_resumo["totais"]["num_itens"] == 4
    assert projeto_resumo["totais"]["receita_mensal"] == 10000
    assert "investimento" in projeto_resumo["graficos"]


def test_dashboard_do_projeto_sem_itens(cliente, admin):
    projeto = _criar_projeto(cliente)
    _criar_local(cliente, projeto["id"], valor_mensal=0, mao_de_obra=0)

    resposta = cliente.get(f"/api/projetos/{projeto['id']}")

    assert resposta.status_code == 200
    assert resposta.json()["locais"][0]["itens"] == []


def test_atualizar_local_recalcula(cliente, admin):
    projeto = _criar_projeto(cliente)
    local = _criar_local(cliente, projeto["id"])
    resposta = cliente.patch(
        f"/api/projetos/{projeto['id']}/locais/{local['id']}",
        json={"valor_mensal": 20000, "nome": "LOCAL B"},
    )
    assert resposta.status_code == 200
    dados = cliente.get(f"/api/projetos/{projeto['id']}").json()
    resumo = dados["locais"][0]["resumo"]
    assert resumo["local"] == "LOCAL B"
    assert resumo["impostos"] == 3000
    assert resumo["saldo_mensal"] == 17000


def test_item_valor_total_automatico(cliente, admin):
    projeto = _criar_projeto(cliente)
    local = _criar_local(cliente, projeto["id"])
    item = _criar_item(cliente, local["id"], qtd=5, valor_unit=120)
    assert item["valor_total"] == 600

    resposta = cliente.patch(f"/api/projetos/itens/{item['id']}", json={"qtd": 10})
    assert resposta.status_code == 200
    assert resposta.json()["valor_total"] == 1200


def test_importar_planilha(cliente, admin):
    resposta = cliente.post(
        "/api/projetos/importar",
        files={"arquivo": ("planilha.xlsx", planilha_base().getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resposta.status_code == 200, resposta.text
    criado = resposta.json()
    assert criado["nome"] == "planilha"

    dados = cliente.get(f"/api/projetos/{criado['id']}").json()
    assert len(dados["locais"]) == 2
    sesc = next(l for l in dados["locais"] if l["nome"] == "SESC TESTE")
    assert sesc["resumo"]["valor_mensal"] == 10000
    assert sesc["resumo"]["impostos"] == 1500
    assert sesc["resumo"]["saldo_mensal"] == 8500
    assert sesc["resumo"]["investimento"] == 2000 + 1500 + 8 * 120 + 4 * 850 + 1400
    assert len(sesc["itens"]) == 4


def test_importar_arquivo_invalido(cliente, admin):
    resposta = cliente.post(
        "/api/projetos/importar",
        files={"arquivo": ("p.xlsx", b"nao-e-xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resposta.status_code == 400


def test_exportar_planilha(cliente, admin, tmp_path):
    projeto = _criar_projeto(cliente)
    local = _criar_local(cliente, projeto["id"])
    _criar_item(cliente, local["id"], cod="A1")
    _criar_item(cliente, local["id"], categoria="CFTV", material="Câmera", qtd=2, valor_unit=850, cod="C1")

    resposta = cliente.get(f"/api/projetos/{projeto['id']}/planilha.xlsx")
    assert resposta.status_code == 200
    assert resposta.content[:2] == b"PK"

    caminho = tmp_path / "planilha.xlsx"
    caminho.write_bytes(resposta.content)
    wb = openpyxl.load_workbook(caminho)
    assert "RELATORIO" in wb.sheetnames
    assert "SESC TESTE" in wb.sheetnames
    assert "INSIGHTS" in wb.sheetnames
    ws_rel = wb["RELATORIO"]
    assert ws_rel["A1"].value == "LOCAL"
    assert ws_rel["B2"].value == 10000
    assert ws_rel["M2"].value == 2000 + 1500 + 2 * 850
    ws_local = wb["SESC TESTE"]
    valores = [ws_local.cell(row=r, column=2).value for r in range(1, ws_local.max_row + 1)]
    assert "MATERIAL ALARME" in valores
    assert "MATERIAL CFTV" in valores
    ws_insights = wb["INSIGHTS"]
    assert [ws_insights.cell(row=1, column=coluna).value for coluna in range(1, 4)] == [
        "LOCAL", "SEVERIDADE", "INSIGHT"
    ]
    assert ws_insights["A2"].value == "SESC TESTE"
    assert ws_insights["B2"].value in {"ok", "dica", "atencao", "alerta"}
    assert "SESC TESTE" in ws_insights["C2"].value


def test_relatorio_pdf_do_projeto(cliente, admin):
    projeto = _criar_projeto(cliente)
    _criar_local(cliente, projeto["id"])
    resposta = cliente.post(f"/api/projetos/{projeto['id']}/relatorio")
    assert resposta.status_code == 200
    assert resposta.content[:4] == b"%PDF"


def test_relatorio_sem_locais(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.post(f"/api/projetos/{projeto['id']}/relatorio")
    assert resposta.status_code == 400


def test_admin_limita_usuarios(cliente, admin):
    cliente.post("/api/users", json={"nome": "A", "username": "a", "senha": "senha12345", "papel": "admin"})
    cliente.post("/api/users", json={"nome": "B", "username": "b", "senha": "senha12345", "papel": "admin"})
    resposta = cliente.post(
        "/api/users", json={"nome": "C", "username": "c", "senha": "senha12345", "papel": "admin"}
    )
    assert resposta.status_code == 400


def test_usuario_comum_sem_acesso_admin(cliente, admin):
    cliente.post(
        "/api/users", json={"nome": "Comum", "username": "comum", "senha": "senha12345", "papel": "usuario"}
    )
    cliente.post("/api/auth/logout")
    _login(cliente, username="comum", senha="senha12345")
    assert cliente.get("/api/users").status_code == 403
    assert cliente.get("/api/projetos").status_code == 200


def _criar_cliente(contexto_cliente, username="cliente1", nome="Cliente 1") -> int:
    resposta = contexto_cliente.post(
        "/api/users",
        json={"nome": nome, "username": username, "senha": "cliente1234", "papel": "cliente"},
    )
    assert resposta.status_code == 200, resposta.text
    return int(resposta.json()["id"])


def _login_como(cliente, username, senha):
    cliente.post("/api/auth/logout")
    resposta = cliente.post("/api/auth/login", json={"username": username, "senha": senha})
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def test_admin_cria_cliente(cliente, admin):
    resposta = cliente.post(
        "/api/users",
        json={"nome": "Cliente X", "username": "cliente_x", "senha": "cliente1234", "papel": "cliente"},
    )
    assert resposta.status_code == 200, resposta.text
    corpo = resposta.json()
    assert corpo["papel"] == "cliente"
    assert corpo["username"] == "cliente_x"


def test_papel_invalido_rejeitado(cliente, admin):
    resposta = cliente.post(
        "/api/users",
        json={"nome": "Z", "username": "z", "senha": "senha12345", "papel": "visitante"},
    )
    assert resposta.status_code == 400


def test_cliente_ve_apenas_seu_projeto(cliente, admin):
    id_cliente_a = _criar_cliente(cliente, username="cliente_a", nome="Cliente A")
    id_cliente_b = _criar_cliente(cliente, username="cliente_b", nome="Cliente B")
    projeto_a = _criar_projeto(cliente, nome="Proj A")
    projeto_b = _criar_projeto(cliente, nome="Proj B")
    cliente.patch(f"/api/projetos/{projeto_a['id']}", json={"cliente_usuario_id": id_cliente_a})
    cliente.patch(f"/api/projetos/{projeto_b['id']}", json={"cliente_usuario_id": id_cliente_b})

    _login_como(cliente, "cliente_a", "cliente1234")
    lista = cliente.get("/api/projetos").json()
    assert len(lista) == 1
    assert lista[0]["id"] == projeto_a["id"]


def test_cliente_nao_acessa_projeto_de_outro(cliente, admin):
    _criar_cliente(cliente, username="cliente_a", nome="Cliente A")
    _criar_cliente(cliente, username="cliente_b", nome="Cliente B")
    projeto_a = _criar_projeto(cliente, nome="Proj A")
    _criar_projeto(cliente, nome="Proj B")
    cliente.patch(f"/api/projetos/{projeto_a['id']}", json={"cliente_usuario_id": 2})

    _login_como(cliente, "cliente_b", "cliente1234")
    assert cliente.get(f"/api/projetos/{projeto_a['id']}").status_code == 404


def test_cliente_pode_baixar_pdf_e_planilha(cliente, admin):
    id_cliente = _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    projeto = _criar_projeto(cliente, nome="Proj X")
    cliente.patch(f"/api/projetos/{projeto['id']}", json={"cliente_usuario_id": id_cliente})
    _criar_local(cliente, projeto["id"])

    _login_como(cliente, "cliente_x", "cliente1234")
    assert cliente.get(f"/api/projetos/{projeto['id']}/planilha.xlsx").status_code == 200
    assert cliente.post(f"/api/projetos/{projeto['id']}/relatorio").status_code == 200


def test_cliente_nao_pode_criar_projeto(cliente, admin):
    _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    _login_como(cliente, "cliente_x", "cliente1234")
    resposta = cliente.post("/api/projetos", json={"nome": "Inv"})
    assert resposta.status_code == 403


def test_cliente_nao_pode_editar_local(cliente, admin):
    id_cliente = _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    projeto = _criar_projeto(cliente, nome="Proj X")
    cliente.patch(f"/api/projetos/{projeto['id']}", json={"cliente_usuario_id": id_cliente})
    local = _criar_local(cliente, projeto["id"])

    _login_como(cliente, "cliente_x", "cliente1234")
    resposta = cliente.patch(
        f"/api/projetos/{projeto['id']}/locais/{local['id']}",
        json={"valor_mensal": 999},
    )
    assert resposta.status_code == 403


def test_cliente_nao_pode_excluir_projeto(cliente, admin):
    id_cliente = _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    projeto = _criar_projeto(cliente, nome="Proj X")
    cliente.patch(f"/api/projetos/{projeto['id']}", json={"cliente_usuario_id": id_cliente})

    _login_como(cliente, "cliente_x", "cliente1234")
    assert cliente.delete(f"/api/projetos/{projeto['id']}").status_code == 403


def test_cliente_nao_pode_importar_planilha(cliente, admin):
    from fixtures import planilha_base

    _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    _login_como(cliente, "cliente_x", "cliente1234")
    resposta = cliente.post(
        "/api/projetos/importar",
        files={"arquivo": ("planilha.xlsx", planilha_base().getvalue(),
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resposta.status_code == 403


def test_cliente_nao_pode_listar_usuarios(cliente, admin):
    _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    _login_como(cliente, "cliente_x", "cliente1234")
    assert cliente.get("/api/users").status_code == 403


def test_patch_nao_altera_cliente_quando_ausente(cliente, admin):
    id_cliente = _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    projeto = _criar_projeto(cliente, nome="Proj X")
    cliente.patch(f"/api/projetos/{projeto['id']}", json={"cliente_usuario_id": id_cliente})
    resposta = cliente.patch(f"/api/projetos/{projeto['id']}", json={"nome": "Renomeado"})
    assert resposta.status_code == 200
    assert resposta.json()["cliente_usuario_id"] == id_cliente
    assert resposta.json()["nome"] == "Renomeado"


def test_patch_atribui_cliente_nulo(cliente, admin):
    id_cliente = _criar_cliente(cliente, username="cliente_x", nome="Cliente X")
    projeto = _criar_projeto(cliente, nome="Proj X")
    cliente.patch(f"/api/projetos/{projeto['id']}", json={"cliente_usuario_id": id_cliente})
    resposta = cliente.patch(f"/api/projetos/{projeto['id']}", json={"cliente_usuario_id": None})
    assert resposta.status_code == 200
    assert resposta.json()["cliente_usuario_id"] is None
