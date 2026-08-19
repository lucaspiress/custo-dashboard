import csv
import io

import openpyxl

import datasets_store


def _login(cliente, username="admin", senha="admin123456"):
    resposta = cliente.post("/api/auth/login", json={"username": username, "senha": senha})
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def _csv_bytes(texto: str) -> bytes:
    return texto.encode("utf-8")


def _xlsx_bytes(linhas: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for linha in linhas:
        ws.append(linha)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _criar_projeto(cliente, nome="Projeto Datasets"):
    resposta = cliente.post("/api/projetos", json={"nome": nome})
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def _criar_local(cliente, projeto_id, **campos):
    dados = {"nome": "SESC TESTE", "valor_mensal": 10000, "mao_de_obra": 2000}
    dados.update(campos)
    resposta = cliente.post(f"/api/projetos/{projeto_id}/locais", json=dados)
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def _criar_item(cliente, local_id, **campos):
    dados = {"categoria": "ALARME", "material": "Central de alarme", "qtd": 1, "valor_unit": 1500}
    dados.update(campos)
    resposta = cliente.post(f"/api/projetos/locais/{local_id}/itens", json=dados)
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def _criar_dataset(cliente, projeto_id, nome="Dataset Livre", schema_json=None):
    corpo = {"nome": nome}
    if schema_json is not None:
        corpo["schema_json"] = schema_json
    resposta = cliente.post(f"/api/projetos/{projeto_id}/datasets", json=corpo)
    assert resposta.status_code == 200, resposta.text
    return resposta.json()


def test_criar_dataset_livre(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"], nome="Meu Dataset")
    assert dataset["id"] > 0
    assert dataset["nome"] == "Meu Dataset"
    assert dataset["fonte"] == "livre"
    assert dataset["projeto_id"] == projeto["id"]
    assert dataset["schema_json"] == {}


def test_criar_dataset_sem_nome(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.post(f"/api/projetos/{projeto['id']}/datasets", json={"nome": "  "})
    assert resposta.status_code == 400


def test_limite_20_datasets(cliente, admin):
    projeto = _criar_projeto(cliente)
    for i in range(20):
        _criar_dataset(cliente, projeto["id"], nome=f"Dataset {i}")
    resposta = cliente.post(
        f"/api/projetos/{projeto['id']}/datasets",
        json={"nome": "Dataset 21"},
    )
    assert resposta.status_code == 400
    assert "20" in resposta.json()["detail"]


def test_listar_datasets_inclui_virtuais(cliente, admin):
    projeto = _criar_projeto(cliente)
    _criar_dataset(cliente, projeto["id"], nome="Livro")
    lista = cliente.get(f"/api/projetos/{projeto['id']}/datasets").json()
    ids = [d["id"] for d in lista]
    assert f"locais-{projeto['id']}" in ids
    assert f"itens-{projeto['id']}" in ids
    assert len(lista) == 3
    virtuais = {d["id"]: d for d in lista if isinstance(d["id"], str)}
    assert virtuais[f"locais-{projeto['id']}"]["fonte"] == "locais"
    assert virtuais[f"itens-{projeto['id']}"]["fonte"] == "itens"
    assert len(virtuais[f"locais-{projeto['id']}"]["schema_json"]) == 11
    assert len(virtuais[f"itens-{projeto['id']}"]["schema_json"]) == 8


def test_obter_dataset_virtual(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.get(f"/api/projetos/{projeto['id']}/datasets/locais-{projeto['id']}")
    assert resposta.status_code == 200
    assert resposta.json()["fonte"] == "locais"


def test_adicionar_linhas(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    resposta = cliente.post(
        f"/api/datasets/{dataset['id']}/rows",
        json={"rows": [
            {"row_index": 0, "data_json": {"nome": "A", "valor": 10}},
            {"row_index": 1, "data_json": {"nome": "B", "valor": 20}},
        ]},
    )
    assert resposta.status_code == 200
    assert resposta.json()["adicionadas"] == 2
    linhas = cliente.get(f"/api/datasets/{dataset['id']}/rows").json()
    assert len(linhas) == 2
    assert linhas[0]["data_json"]["nome"] == "A"
    assert linhas[1]["data_json"]["valor"] == 20


def test_adicionar_linhas_upsert(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    cliente.post(
        f"/api/datasets/{dataset['id']}/rows",
        json={"rows": [{"row_index": 0, "data_json": {"v": 1}}]},
    )
    resposta = cliente.post(
        f"/api/datasets/{dataset['id']}/rows",
        json={"rows": [{"row_index": 0, "data_json": {"v": 2}}]},
    )
    assert resposta.status_code == 200
    linhas = cliente.get(f"/api/datasets/{dataset['id']}/rows").json()
    assert len(linhas) == 1
    assert linhas[0]["data_json"]["v"] == 2


def test_adicionar_linhas_excede_100k(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    linhas = [{"row_index": i, "data_json": {"v": i}} for i in range(100000)]
    datasets_store.adicionar_linhas(dataset["id"], projeto["id"], linhas)
    resposta = cliente.post(
        f"/api/datasets/{dataset['id']}/rows",
        json={"rows": [{"row_index": 100000, "data_json": {"v": 1}}]},
    )
    assert resposta.status_code == 400
    assert "100000" in resposta.json()["detail"]


def test_get_locais_virtual(cliente, admin):
    projeto = _criar_projeto(cliente)
    _criar_local(cliente, projeto["id"], nome="LOCAL A")
    _criar_local(cliente, projeto["id"], nome="LOCAL B")
    resposta = cliente.get(f"/api/datasets/locais-{projeto['id']}/rows")
    assert resposta.status_code == 200
    linhas = resposta.json()
    assert len(linhas) == 2
    nomes = [l["data_json"]["nome"] for l in linhas]
    assert "LOCAL A" in nomes
    assert "LOCAL B" in nomes


def test_get_itens_virtual(cliente, admin):
    projeto = _criar_projeto(cliente)
    local = _criar_local(cliente, projeto["id"])
    _criar_item(cliente, local["id"], material="Central de alarme")
    _criar_item(cliente, local["id"], material="Câmera IP", categoria="CFTV")
    resposta = cliente.get(f"/api/datasets/itens-{projeto['id']}/rows")
    assert resposta.status_code == 200
    linhas = resposta.json()
    assert len(linhas) == 2
    materiais = [l["data_json"]["material"] for l in linhas]
    assert "Central de alarme" in materiais
    assert "Câmera IP" in materiais


def test_patch_virtual_405(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.patch(
        f"/api/projetos/{projeto['id']}/datasets/locais-{projeto['id']}",
        json={"nome": "X"},
    )
    assert resposta.status_code == 405


def test_delete_virtual_405(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.delete(f"/api/projetos/{projeto['id']}/datasets/itens-{projeto['id']}")
    assert resposta.status_code == 405


def test_post_rows_virtual_405(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.post(
        f"/api/datasets/locais-{projeto['id']}/rows",
        json={"rows": [{"row_index": 0, "data_json": {}}]},
    )
    assert resposta.status_code == 405


def test_atualizar_dataset(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"], nome="Antes")
    resposta = cliente.patch(
        f"/api/projetos/{projeto['id']}/datasets/{dataset['id']}",
        json={"nome": "Depois", "schema_json": {"colunas": [{"campo": "x", "tipo": "text"}]}},
    )
    assert resposta.status_code == 200
    assert resposta.json()["nome"] == "Depois"
    assert resposta.json()["schema_json"]["colunas"][0]["campo"] == "x"


def test_delete_dataset_cascata(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    cliente.post(
        f"/api/datasets/{dataset['id']}/rows",
        json={"rows": [{"row_index": 0, "data_json": {"v": 1}}]},
    )
    assert cliente.delete(f"/api/projetos/{projeto['id']}/datasets/{dataset['id']}").status_code == 204
    assert cliente.get(f"/api/projetos/{projeto['id']}/datasets/{dataset['id']}").status_code == 404
    assert cliente.get(f"/api/datasets/{dataset['id']}/rows").status_code == 404


def test_sem_login_401(cliente):
    assert cliente.get("/api/projetos/1/datasets").status_code == 401
    assert cliente.post("/api/projetos/1/datasets", json={"nome": "X"}).status_code == 401
    assert cliente.get("/api/projetos/1/datasets/locais-1").status_code == 401
    assert cliente.patch("/api/projetos/1/datasets/locais-1", json={"nome": "X"}).status_code == 401
    assert cliente.delete("/api/projetos/1/datasets/locais-1").status_code == 401
    assert cliente.get("/api/datasets/1/rows").status_code == 401
    assert cliente.post("/api/datasets/1/rows", json={"rows": []}).status_code == 401


# ---------------------------------------------------------------- import/export

def test_importar_csv_simples(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    resposta = cliente.post(
        f"/api/datasets/{dataset['id']}/importar",
        files={"arquivo": ("dados.csv", _csv_bytes("nome,valor\nA,10\nB,20\n"), "text/csv")},
    )
    assert resposta.status_code == 200, resposta.text
    corpo = resposta.json()
    assert corpo["colunas"] == ["nome", "valor"]
    assert corpo["linhas_adicionadas"] == 2
    assert corpo["tipos"]["nome"] == "text"
    assert corpo["tipos"]["valor"] == "number"
    linhas = cliente.get(f"/api/datasets/{dataset['id']}/rows").json()
    assert len(linhas) == 2
    assert linhas[0]["data_json"]["nome"] == "A"


def test_importar_csv_formato_br(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    resposta = cliente.post(
        f"/api/datasets/{dataset['id']}/importar",
        files={"arquivo": ("dados.csv", _csv_bytes("produto;preco\nX;1.234,56\nY;99,90\n"), "text/csv")},
    )
    assert resposta.status_code == 200, resposta.text
    corpo = resposta.json()
    assert corpo["tipos"]["preco"] == "number"
    linhas = cliente.get(f"/api/datasets/{dataset['id']}/rows").json()
    precos = {l["data_json"]["produto"]: l["data_json"]["preco"] for l in linhas}
    assert precos["X"] == "1234.56"
    assert precos["Y"] == "99.90"


def test_importar_xlsx(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    conteudo = _xlsx_bytes([
        ["nome", "qtd", "valor"],
        ["A", 2, 10.5],
        ["B", 3, 20.0],
    ])
    resposta = cliente.post(
        f"/api/datasets/{dataset['id']}/importar",
        files={"arquivo": ("dados.xlsx", conteudo,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resposta.status_code == 200, resposta.text
    corpo = resposta.json()
    assert corpo["colunas"] == ["nome", "qtd", "valor"]
    assert corpo["linhas_adicionadas"] == 2
    assert corpo["tipos"]["nome"] == "text"
    assert corpo["tipos"]["qtd"] == "number"
    assert corpo["tipos"]["valor"] == "number"


def test_importar_arquivo_grande_413(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    grande = b"x" * (10 * 1024 * 1024 + 1)
    resposta = cliente.post(
        f"/api/datasets/{dataset['id']}/importar",
        files={"arquivo": ("grande.csv", grande, "text/csv")},
    )
    assert resposta.status_code == 413


def test_importar_virtual_405(cliente, admin):
    projeto = _criar_projeto(cliente)
    resposta = cliente.post(
        f"/api/datasets/locais-{projeto['id']}/importar",
        files={"arquivo": ("d.csv", _csv_bytes("a\n1\n"), "text/csv")},
    )
    assert resposta.status_code == 405


def test_exportar_csv_dataset_livre(cliente, admin):
    projeto = _criar_projeto(cliente)
    dataset = _criar_dataset(cliente, projeto["id"])
    cliente.post(
        f"/api/datasets/{dataset['id']}/importar",
        files={"arquivo": ("d.csv", _csv_bytes("nome,valor\nA,10\nB,20\n"), "text/csv")},
    )
    resposta = cliente.get(f"/api/datasets/{dataset['id']}/export.csv")
    assert resposta.status_code == 200
    assert resposta.headers["content-type"].startswith("text/csv")
    texto = resposta.content.decode("utf-8")
    linhas = texto.strip().splitlines()
    assert linhas[0] == "nome,valor"
    assert "A" in linhas[1]
    assert "B" in linhas[2]


def test_exportar_csv_locais_virtual(cliente, admin):
    projeto = _criar_projeto(cliente)
    _criar_local(cliente, projeto["id"], nome="LOCAL A")
    _criar_local(cliente, projeto["id"], nome="LOCAL B")
    resposta = cliente.get(f"/api/datasets/locais-{projeto['id']}/export.csv")
    assert resposta.status_code == 200
    texto = resposta.content.decode("utf-8")
    assert "nome" in texto.splitlines()[0]
    assert "LOCAL A" in texto
    assert "LOCAL B" in texto


def test_exportar_xlsx_itens_virtual(cliente, admin):
    projeto = _criar_projeto(cliente)
    local = _criar_local(cliente, projeto["id"])
    _criar_item(cliente, local["id"], material="Central de alarme")
    resposta = cliente.get(f"/api/datasets/itens-{projeto['id']}/export.xlsx")
    assert resposta.status_code == 200
    assert resposta.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(resposta.content))
    ws = wb.active
    cabecalho = [c.value for c in ws[1]]
    assert "material" in cabecalho
    idx_material = cabecalho.index("material")
    assert ws.cell(row=2, column=idx_material + 1).value == "Central de alarme"
