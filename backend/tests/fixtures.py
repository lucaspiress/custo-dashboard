from io import BytesIO

import openpyxl

SHEET_RELATORIO = "RELATORIO"
SHEET_GRAFICOS = "GRÁFICOS"


def _montar_aba_local(ws, categoria: str, itens: list[dict]) -> None:
    ws.append([None, f"MATERIAL {categoria}"])
    for item in itens:
        ws.append(
            [
                item["cod"],
                item["material"],
                item["qtd"],
                item["valor_unit"],
                item["qtd"] * item["valor_unit"],
            ]
        )
    ws.append(["TOTAL"])


def montar_planilha(
    locais: list[dict],
    meses: int = 12,
) -> BytesIO:
    wb = openpyxl.Workbook()
    ws_rel = wb.active
    ws_rel.title = SHEET_RELATORIO

    for local in locais:
        nome = local["nome"]
        wb.create_sheet(nome)
        ws_local = wb[nome]
        if "itens" in local:
            for bloco in local["itens"]:
                _montar_aba_local(ws_local, bloco["categoria"], bloco["itens"])

    ws_rel.append(["LOCAL", "VALOR MENSAL", "TAXA INSTALACAO", "IMPOSTOS", "SALDO APOS",
                   "CUSTO MANUT", "TERCEIRIZADA", "CHIP", "SOFTWARES", "SALDO MENSAL",
                   "MAO DE OBRA", "EQUIPAMENTO", "INVESTIMENTO", "TEMPO RETORNO", "DATA INST"])
    for linha, local in enumerate(locais, start=2):
        nome = local["nome"]
        equipamento = sum(
            item["qtd"] * item["valor_unit"]
            for bloco in local.get("itens", [])
            for item in bloco["itens"]
        )
        ws_rel.cell(row=linha, column=1, value=nome)
        ws_rel.cell(row=linha, column=2, value=local.get("valor_mensal", 5000))
        ws_rel.cell(row=linha, column=3, value=local.get("taxa_instalacao", 0))
        ws_rel.cell(row=linha, column=6, value=local.get("custo_manutencao", 0))
        ws_rel.cell(row=linha, column=7, value=local.get("mensal_terceirizada", 0))
        ws_rel.cell(row=linha, column=8, value=local.get("chip_mensal", 0))
        ws_rel.cell(row=linha, column=9, value=local.get("custos_softwares", 0))
        ws_rel.cell(row=linha, column=11, value=local.get("mao_de_obra", 1000))
        ws_rel.cell(row=linha, column=12, value=f"='{nome}'!E99")
        ws_rel.cell(row=linha, column=13, value=local.get("mao_de_obra", 1000) + equipamento)
        ws_rel.cell(row=linha, column=15, value=local.get("data_inst", "01/08/2026"))

    wb.create_sheet(SHEET_GRAFICOS)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def planilha_base() -> BytesIO:
    return montar_planilha(
        [
            {
                "nome": "SESC TESTE",
                "valor_mensal": 10000,
                "mao_de_obra": 2000,
                "itens": [
                    {
                        "categoria": "ALARME",
                        "itens": [
                            {"cod": "A1", "material": "Central de alarme", "qtd": 1, "valor_unit": 1500},
                            {"cod": "A2", "material": "Sensor de presença", "qtd": 8, "valor_unit": 120},
                        ],
                    },
                    {
                        "categoria": "CFTV",
                        "itens": [
                            {"cod": "C1", "material": "Câmera IP 2MP", "qtd": 4, "valor_unit": 850},
                            {"cod": "C2", "material": "DVR 8 canais", "qtd": 1, "valor_unit": 1400},
                        ],
                    },
                ],
            },
            {
                "nome": "UNIDADE B",
                "valor_mensal": 6000,
                "mao_de_obra": 1200,
                "itens": [
                    {
                        "categoria": "CFTV",
                        "itens": [
                            {"cod": "C1", "material": "Câmera IP 2MP", "qtd": 2, "valor_unit": 850},
                        ],
                    },
                ],
            },
        ]
    )


def projeto_escopo(numero: int) -> dict:
    """Return deterministic project data with unique isolation markers.

    The returned values are request payloads, so tests can use the fixture
    without depending on database-generated IDs.  ``numero`` is deliberately
    reflected in both the names and the numeric values to make accidental
    cross-project reads easy to spot.
    """
    if numero < 1:
        raise ValueError("numero deve ser positivo")

    marcador = f"SC001-P{numero:02d}"
    return {
        "nome": marcador,
        "cliente": f"Cliente {marcador}",
        "local": {
            "nome": f"{marcador}-LOCAL",
            "valor_mensal": 10000 + numero * 1000,
            "taxa_instalacao": 100 + numero,
            "custo_manutencao": 200 + numero,
            "mensal_terceirizada": 300 + numero,
            "chip_mensal": 40 + numero,
            "custos_softwares": 50 + numero,
            "mao_de_obra": 1000 + numero * 10,
        },
        "item": {
            "categoria": f"SC001-CATEGORIA-{numero:02d}",
            "cod": f"{marcador}-COD",
            "material": f"{marcador}-ITEM",
            "qtd": numero,
            "valor_unit": 500 + numero,
        },
    }


def projetos_escopo(quantidade: int = 10) -> list[dict]:
    """Return the deterministic project set used by isolation checks."""
    if quantidade < 1:
        raise ValueError("quantidade deve ser positiva")
    return [projeto_escopo(numero) for numero in range(1, quantidade + 1)]


def criar_projeto_escopo(cliente, numero: int) -> dict:
    """Create one isolated project, local and item through the existing API."""
    fixture = projeto_escopo(numero)

    resposta = cliente.post(
        "/api/projetos",
        json={"nome": fixture["nome"], "cliente": fixture["cliente"]},
    )
    assert resposta.status_code == 200, resposta.text
    projeto = resposta.json()

    resposta = cliente.post(
        f"/api/projetos/{projeto['id']}/locais",
        json=fixture["local"],
    )
    assert resposta.status_code == 200, resposta.text
    local = resposta.json()

    resposta = cliente.post(
        f"/api/projetos/locais/{local['id']}/itens",
        json=fixture["item"],
    )
    assert resposta.status_code == 200, resposta.text
    item = resposta.json()

    return {
        "fixture": fixture,
        "projeto": projeto,
        "local": local,
        "item": item,
    }


def criar_projetos_escopo(cliente, quantidade: int = 10) -> list[dict]:
    """Create a deterministic set of project-scoped API test records."""
    if quantidade < 1:
        raise ValueError("quantidade deve ser positiva")
    return [criar_projeto_escopo(cliente, numero) for numero in range(1, quantidade + 1)]
