import io
from dataclasses import dataclass, field
from datetime import datetime, timedelta

import openpyxl

import config

EXCEL_EPOCH = datetime(1899, 12, 30)


@dataclass
class Item:
    cod: str
    material: str
    qtd: float
    valor_unit: float
    valor_total: float
    categoria: str


@dataclass
class Local:
    nome: str
    valor_mensal: float
    taxa_instalacao: float
    custo_manutencao: float
    mensal_terceirizada: float
    chip_mensal: float
    custos_softwares: float
    mao_de_obra: float
    data_inst: datetime | None
    itens: list[Item] = field(default_factory=list)

    @property
    def impostos(self) -> float:
        return self.valor_mensal * config.TAXA_IMPOSTOS

    @property
    def saldo_apos_impostos(self) -> float:
        return self.valor_mensal - self.impostos

    @property
    def custos_fixos(self) -> float:
        return (
            self.custo_manutencao
            + self.mensal_terceirizada
            + self.chip_mensal
            + self.custos_softwares
        )

    @property
    def saldo_mensal(self) -> float:
        return self.saldo_apos_impostos - self.custos_fixos

    @property
    def equipamento(self) -> float:
        return sum(item.valor_total for item in self.itens)

    @property
    def investimento(self) -> float:
        return self.mao_de_obra + self.equipamento

    @property
    def receita_anual(self) -> float:
        return (self.valor_mensal * config.MESES_POR_ANO) + self.taxa_instalacao

    @property
    def tempo_retorno(self) -> float | None:
        if self.saldo_mensal <= 0:
            return None
        return (self.investimento - self.taxa_instalacao) / self.saldo_mensal

    @property
    def margem(self) -> float | None:
        if self.valor_mensal <= 0:
            return None
        return self.saldo_mensal / self.valor_mensal


@dataclass
class WorkbookData:
    locais: list[Local]
    avisos: list[str] = field(default_factory=list)


def _to_float(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace("R$", "").strip()
    if not texto or texto == "-":
        return 0.0
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _to_date(valor) -> datetime | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, (int, float)):
        return EXCEL_EPOCH + timedelta(days=float(valor))
    texto = str(valor).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    return None


def _encontrar_linha_cabecalho(ws) -> int:
    for linha in range(1, min(ws.max_row, 10) + 1):
        valor = ws[f"A{linha}"].value
        if valor is not None and str(valor).strip().upper() == config.HEADER_LOCAL:
            return linha
    raise ValueError(
        "Planilha não reconhecida: não encontrei a coluna 'LOCAL' na aba "
        f"'{config.SHEET_RELATORIO}'. Verifique se o arquivo segue o template."
    )


def _parse_itens(ws, nome_aba: str) -> list[Item]:
    itens: list[Item] = []
    for linha in range(1, ws.max_row + 1):
        valor_b = ws[f"B{linha}"].value
        if valor_b is None:
            continue
        texto_b = str(valor_b).strip()
        if texto_b.upper().startswith(config.PREFIXO_MATERIAL):
            categoria = texto_b[len(config.PREFIXO_MATERIAL):].strip()
            linha_item = linha + 1
            while linha_item <= ws.max_row:
                cod = ws[f"A{linha_item}"].value
                material = ws[f"B{linha_item}"].value
                if cod is not None and str(cod).strip().upper() == config.HEADER_TOTAL:
                    break
                if cod is None or str(cod).strip() == "":
                    break
                if material is None or str(material).strip() == "":
                    break
                qtd = _to_float(ws[f"C{linha_item}"].value)
                valor_unit = _to_float(ws[f"D{linha_item}"].value)
                valor_total_cell = ws[f"E{linha_item}"].value
                if isinstance(valor_total_cell, (int, float)):
                    valor_total = float(valor_total_cell)
                else:
                    valor_total = qtd * valor_unit
                itens.append(
                    Item(
                        cod=str(cod).strip(),
                        material=str(material).strip(),
                        qtd=qtd,
                        valor_unit=valor_unit,
                        valor_total=valor_total,
                        categoria=categoria,
                    )
                )
                linha_item += 1
    return itens


def _aba_para_local(ws_names: list[str], nome_local: str) -> str | None:
    nome_local_norm = nome_local.strip().lower()
    melhores = []
    for nome in ws_names:
        if nome.strip().lower() in (config.SHEET_RELATORIO.lower(), config.SHEET_GRAFICOS.lower()):
            continue
        nome_norm = nome.strip().lower()
        if nome_local_norm.startswith(nome_norm) or nome_norm.startswith(nome_local_norm) or nome_norm in nome_local_norm:
            melhores.append((len(nome_norm), nome))
    if not melhores:
        return None
    melhores.sort(key=lambda x: x[0], reverse=True)
    return melhores[0][1]


def carregar(origem) -> WorkbookData:
    if isinstance(origem, (str, bytes, bytearray)):
        import os
        if isinstance(origem, (bytes, bytearray)):
            wb = openpyxl.load_workbook(io.BytesIO(origem), data_only=False)
        else:
            wb = openpyxl.load_workbook(origem, data_only=False)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(origem.getvalue()), data_only=False)

    nome_abas = [ws.title for ws in wb.worksheets]
    if config.SHEET_RELATORIO not in nome_abas:
        raise ValueError(
            f"Planilha não reconhecida: aba '{config.SHEET_RELATORIO}' não encontrada. "
            "Verifique se o arquivo segue o template."
        )

    ws_relatorio = wb[config.SHEET_RELATORIO]
    linha_cab = _encontrar_linha_cabecalho(ws_relatorio)
    mapa = config.RELATORIO_COLUNAS

    locais: list[Local] = []
    avisos: list[str] = []
    linha = linha_cab + 1
    while linha <= ws_relatorio.max_row:
        nome_cell = ws_relatorio[f"{mapa['local']}{linha}"].value
        if nome_cell is None or str(nome_cell).strip() == "":
            break
        nome = str(nome_cell).strip()
        if nome.upper() == config.HEADER_TOTAL:
            break
        data_inst = _to_date(ws_relatorio[f"{mapa['data_inst']}{linha}"].value)
        local = Local(
            nome=nome,
            valor_mensal=_to_float(ws_relatorio[f"{mapa['valor_mensal']}{linha}"].value),
            taxa_instalacao=_to_float(ws_relatorio[f"{mapa['taxa_instalacao']}{linha}"].value),
            custo_manutencao=_to_float(ws_relatorio[f"{mapa['custo_manutencao']}{linha}"].value),
            mensal_terceirizada=_to_float(ws_relatorio[f"{mapa['mensal_terceirizada']}{linha}"].value),
            chip_mensal=_to_float(ws_relatorio[f"{mapa['chip_mensal']}{linha}"].value),
            custos_softwares=_to_float(ws_relatorio[f"{mapa['custos_softwares']}{linha}"].value),
            mao_de_obra=_to_float(ws_relatorio[f"{mapa['mao_de_obra']}{linha}"].value),
            data_inst=data_inst,
        )
        aba_itens = _aba_para_local(nome_abas, nome)
        if aba_itens is None:
            avisos.append(f"Local '{nome}': nenhuma aba de equipamento encontrada (itens zerados).")
        else:
            local.itens = _parse_itens(wb[aba_itens], aba_itens)
        locais.append(local)
        linha += 1

    if not locais:
        raise ValueError(
            f"Nenhum local encontrado abaixo do cabeçalho da aba '{config.SHEET_RELATORIO}'."
        )

    return WorkbookData(locais=locais, avisos=avisos)
